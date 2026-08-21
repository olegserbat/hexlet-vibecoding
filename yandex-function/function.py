"""Yandex Cloud Function: принимает заявку из Яндекс Форм и дописывает
строку в xlsx-таблицу на Яндекс Диске через Disk REST API.

Переменные окружения (задаются в консоли Cloud Functions, не в коде):
  YANDEX_DISK_TOKEN  — OAuth-токен с доступом на чтение и запись Диска
  DISK_FILE_PATH     — путь к файлу на Диске, например
                        "/Запись на прием (тест).xlsx"
  WEBHOOK_SECRET      — произвольная строка; действие в Формах должно
                        передавать её же в заголовке X-Webhook-Secret
                        или параметре ?secret=, иначе запрос отклоняется
"""

import base64
import io
import json
import os
from datetime import datetime, timedelta, timezone

import requests
from openpyxl import Workbook, load_workbook

DISK_API = "https://cloud-api.yandex.net/v1/disk/resources"
MOSCOW = timezone(timedelta(hours=3))
COLUMNS = ["Дата и время заявки", "Имя", "Телефон", "Почта", "Дата визита", "Причина обращения"]


def _disk_headers():
    return {"Authorization": f"OAuth {os.environ['YANDEX_DISK_TOKEN']}"}


def _file_path():
    return os.environ.get("DISK_FILE_PATH", "/Запись на прием (тест).xlsx")


def _get_header(headers, name):
    for key, value in (headers or {}).items():
        if key.lower() == name.lower():
            return value
    return None


def _check_secret(event):
    expected = os.environ.get("WEBHOOK_SECRET")
    if not expected:
        return True
    provided = _get_header(event.get("headers"), "X-Webhook-Secret")
    if not provided:
        provided = (event.get("queryStringParameters") or {}).get("secret")
    return provided == expected


def _parse_body(event):
    body = event.get("body") or "{}"
    if event.get("isBase64Encoded"):
        body = base64.b64decode(body).decode("utf-8")
    return json.loads(body)


def _download_workbook():
    resp = requests.get(
        f"{DISK_API}/download", headers=_disk_headers(), params={"path": _file_path()}, timeout=10
    )
    if resp.status_code == 404:
        return Workbook()
    resp.raise_for_status()
    file_resp = requests.get(resp.json()["href"], timeout=10)
    file_resp.raise_for_status()
    return load_workbook(io.BytesIO(file_resp.content))


def _upload_workbook(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = requests.get(
        f"{DISK_API}/upload",
        headers=_disk_headers(),
        params={"path": _file_path(), "overwrite": "true"},
        timeout=10,
    )
    resp.raise_for_status()
    put_resp = requests.put(resp.json()["href"], data=buf, timeout=15)
    put_resp.raise_for_status()


def _response(status, payload):
    return {"statusCode": status, "headers": {"Content-Type": "application/json"}, "body": json.dumps(payload)}


def handler(event, context):
    if not _check_secret(event):
        return _response(403, {"error": "forbidden"})

    try:
        data = _parse_body(event)
    except (ValueError, TypeError):
        return _response(400, {"error": "bad json"})

    try:
        wb = _download_workbook()
    except requests.HTTPError:
        return _response(502, {"error": "disk download failed"})

    ws = wb.active
    if ws.max_row == 1 and ws["A1"].value is None:
        # На новом Workbook() append() не попадает в первую строку — она
        # считается уже «занятой», и данные съезжают на строку ниже.
        # Поэтому заголовок пишем напрямую в ячейки, а не через append().
        for col_idx, title in enumerate(COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=title)

    ws.append(
        [
            datetime.now(MOSCOW).strftime("%d.%m.%Y %H:%M"),
            data.get("name", ""),
            data.get("phone", ""),
            data.get("email", ""),
            data.get("visit_date", ""),
            data.get("reason", ""),
        ]
    )

    try:
        _upload_workbook(wb)
    except requests.HTTPError:
        return _response(502, {"error": "disk upload failed"})

    return _response(200, {"ok": True})
